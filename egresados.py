from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from db import get_connection
from functools import wraps
import psycopg2.extras
from psycopg2 import errors
from io import BytesIO
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

egresados_bp = Blueprint("egresados", __name__)

# =========================
# DECORADOR LOGIN
# =========================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated

# =========================
# LISTAR EGRESADOS
# =========================
@egresados_bp.route("/egresados")
@login_required
def listar_egresados():
    q = request.args.get("q", "").strip()
    carrera = request.args.get("carrera", "")
    estatus = request.args.get("estatus", "")

    conn = get_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = "SELECT * FROM egresados WHERE 1=1"
    params = []

    if q:
        query += " AND (matricula ILIKE %s OR nombre_completo ILIKE %s)"
        params.extend([f"%{q}%", f"%{q}%"])

    if carrera:
        query += " AND carrera = %s"
        params.append(carrera)

    if estatus:
        query += " AND estatus = %s"
        params.append(estatus)

    query += " ORDER BY id DESC"

    cur.execute(query, params)
    egresados = cur.fetchall()

    # Para los selects
    cur.execute("SELECT DISTINCT carrera FROM egresados ORDER BY carrera")
    carreras = [c[0] for c in cur.fetchall()]

    cur.execute("SELECT DISTINCT estatus FROM egresados ORDER BY estatus")
    estatuses = [e[0] for e in cur.fetchall()]

    cur.close()
    conn.close()

    return render_template(
        "egresados.html",
        egresados=egresados,
        carreras=carreras,
        estatuses=estatuses,
        q=q,
        carrera_sel=carrera,
        estatus_sel=estatus
    )

# =========================
# VER EGRESADO (NUEVA RUTA)
# =========================
@egresados_bp.route("/egresados/ver/<int:id>")
@login_required
def ver_egresado(id):
    conn = get_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cur.execute("SELECT * FROM egresados WHERE id = %s", (id,))
    egresado = cur.fetchone()
    
    cur.close()
    conn.close()
    
    if not egresado:
        flash("Egresado no encontrado.", "error")
        return redirect(url_for("egresados.listar_egresados"))
    
    return render_template("ver_egresado.html", egresado=egresado)

# =========================
# REGISTRAR EGRESADO
# =========================
@egresados_bp.route("/egresados/registrar", methods=["GET", "POST"])
@login_required
def registrar_egresado():

    if request.method == "GET":
        return render_template("registrar_egresado.html", registrado=False)

    matricula = request.form["matricula"].strip()

    # Validación matrícula
    if not matricula.isdigit() or len(matricula) != 8:
        flash("La matrícula debe tener exactamente 8 dígitos.", "error")
        return redirect(url_for("egresados.registrar_egresado"))

    # Generación OTRA
    generacion = request.form["generacion"]
    if generacion == "OTRA":
        generacion = request.form.get("otra_generacion", "").strip()
        if not generacion:
            flash("Debe especificar la generación.", "error")
            return redirect(url_for("egresados.registrar_egresado"))

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO egresados
            (matricula, nombre_completo, carrera, generacion, estatus,
             domicilio, genero, telefono, correo)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            matricula,
            request.form["nombre"],
            request.form["carrera"],
            generacion,
            request.form["estatus"],
            request.form.get("domicilio"),
            request.form.get("genero"),
            request.form.get("telefono"),
            request.form.get("correo")
        ))

        conn.commit()

    except errors.UniqueViolation:
        conn.rollback()
        flash("⚠️ La matrícula ya ha sido registrada.", "error")
        return redirect(url_for("egresados.registrar_egresado"))

    finally:
        cur.close()
        conn.close()

    return render_template("registrar_egresado.html", registrado=True)

# =========================
# EDITAR EGRESADO
# =========================
@egresados_bp.route("/egresados/editar/<int:id>", methods=["GET", "POST"])
@login_required
def editar_egresado(id):
    conn = get_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    if request.method == "POST":
        matricula = request.form["matricula"].strip()
        
        # Validar formato de matrícula
        if not matricula.isdigit() or len(matricula) != 8:
            flash("La matrícula debe tener exactamente 8 dígitos.", "error")
            cur.execute("SELECT * FROM egresados WHERE id = %s", (id,))
            egresado = cur.fetchone()
            cur.close()
            conn.close()
            return render_template("editar_egresado.html", egresado=egresado)
        
        # Verificar si la matrícula ya existe en OTRO egresado
        cur.execute("SELECT id FROM egresados WHERE matricula = %s AND id != %s", 
                   (matricula, id))
        if cur.fetchone():
            flash("⚠️ La matrícula ya está registrada en otro egresado.", "error")
            cur.execute("SELECT * FROM egresados WHERE id = %s", (id,))
            egresado = cur.fetchone()
            cur.close()
            conn.close()
            return render_template("editar_egresado.html", egresado=egresado)

        generacion = request.form["generacion"]
        if generacion == "OTRA":
            generacion = request.form.get("otra_generacion", "").strip()

        cur.execute("""
            UPDATE egresados
            SET matricula = %s,
                nombre_completo = %s,
                carrera = %s,
                generacion = %s,
                estatus = %s,
                domicilio = %s,
                genero = %s,
                telefono = %s,
                correo = %s
            WHERE id = %s
        """, (
            matricula,
            request.form["nombre"],
            request.form["carrera"],
            generacion,
            request.form["estatus"],
            request.form.get("domicilio"),
            request.form.get("genero"),
            request.form.get("telefono"),
            request.form.get("correo"),
            id
        ))

        conn.commit()
        flash("✅ Egresado actualizado correctamente.", "success")
        cur.close()
        conn.close()
        return redirect(url_for("egresados.listar_egresados"))

    # GET
    cur.execute("SELECT * FROM egresados WHERE id = %s", (id,))
    egresado = cur.fetchone()
    cur.close()
    conn.close()

    return render_template("editar_egresado.html", egresado=egresado)

# =========================
# ELIMINAR EGRESADO
# =========================
@egresados_bp.route("/egresados/eliminar/<int:id>")
@login_required
def eliminar_egresado(id):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM egresados WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("egresados.listar_egresados"))

@egresados_bp.route("/egresados/exportar/excel")
@login_required
def exportar_excel():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            matricula,
            nombre_completo,
            carrera,
            generacion,
            estatus,
            domicilio,
            genero,
            telefono,
            correo,
            fecha_registro
        FROM egresados
        ORDER BY id
    """)
    datos = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Egresados"

    encabezados = [
        "Matrícula", "Nombre Completo", "Carrera", "Generación",
        "Estatus", "Domicilio", "Género", "Teléfono",
        "Correo", "Fecha de Registro"
    ]

    ws.append(encabezados)

    # 🎨 Estilo encabezados
    from openpyxl.styles import Font, PatternFill, Border, Side

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B3D1E")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, len(encabezados) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # 📄 Datos
    for fila in datos:
        ws.append(fila)

    # 📐 Ajustar ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    cur.close()
    conn.close()

    return send_file(
        buffer,
        as_attachment=True,
        download_name="egresados_completo.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@egresados_bp.route("/egresados/exportar/pdf")
@login_required
def exportar_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from datetime import datetime
    
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            matricula,
            nombre_completo,
            carrera,
            generacion,
            estatus,
            telefono,
            correo
        FROM egresados
        ORDER BY carrera, nombre_completo
    """)
    datos = cur.fetchall()
    
    cur.close()
    conn.close()

    buffer = BytesIO()
    
    # Usar formato horizontal para más espacio
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=40,
        bottomMargin=30
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Estilo para el título
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0B3D1E'),
        spaceAfter=8,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Estilo para subtítulo
    subtitulo_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    # Contenido del documento
    elementos = []
    
    # Título
    titulo = Paragraph("Reporte General de Egresados", titulo_style)
    elementos.append(titulo)
    
    # Fecha de generación
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    subtitulo = Paragraph(f"Generado el: {fecha_actual} | Total de registros: {len(datos)}", subtitulo_style)
    elementos.append(subtitulo)
    
    # Preparar datos para la tabla
    encabezados = [
        ['Matrícula', 'Nombre Completo', 'Carrera', 'Generación', 
         'Estatus', 'Teléfono', 'Correo']
    ]
    
    # Formatear datos con Paragraph para permitir saltos de línea
    datos_tabla = []
    normal_style = ParagraphStyle(
        'CeldaNormal',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
        wordWrap='CJK'
    )
    
    for fila in datos:
        fila_formateada = []
        for i, valor in enumerate(fila):
            texto = str(valor) if valor else 'N/A'
            
            # Columna de carrera (índice 2) - usar Paragraph para saltos automáticos
            if i == 2:
                fila_formateada.append(Paragraph(texto, normal_style))
            # Columnas de nombre y correo también pueden ser largas
            elif i in [1, 6]:
                if len(texto) > 25:
                    fila_formateada.append(Paragraph(texto, normal_style))
                else:
                    fila_formateada.append(texto)
            else:
                # Limitar otras columnas
                if len(texto) > 15:
                    texto = texto[:12] + '...'
                fila_formateada.append(texto)
        
        datos_tabla.append(fila_formateada)
    
    # Combinar encabezados y datos
    tabla_completa = encabezados + datos_tabla
    
    # Crear tabla con anchos de columna optimizados
    tabla = Table(tabla_completa, colWidths=[
        0.8*inch,   # Matrícula
        1.8*inch,   # Nombre
        1.5*inch,   # Carrera
        0.9*inch,   # Generación
        0.9*inch,   # Estatus
        1.0*inch,   # Teléfono
        1.8*inch    # Correo
    ])
    
    # Estilo de la tabla
    estilo_tabla = TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B3D1E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Matrícula centrada
        ('ALIGN', (1, 1), (-1, -1), 'LEFT'),   # Resto a la izquierda
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#0B3D1E')),
        
        # Filas alternadas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        
        # Ajuste de texto
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    
    tabla.setStyle(estilo_tabla)
    elementos.append(tabla)
    
    # Pie de página con info adicional
    elementos.append(Spacer(1, 0.3*inch))
    pie_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    pie = Paragraph(
        "Este documento es un reporte automatizado del sistema de gestión de egresados",
        pie_style
    )
    elementos.append(pie)
    
    # Construir PDF
    doc.build(elementos)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"egresados_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        mimetype="application/pdf"
    )